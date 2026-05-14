from http.server import BaseHTTPRequestHandler
import json, base64, io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import Color


class handler(BaseHTTPRequestHandler):

    def do_GET(self):
        result = {}
        try:
            import openpyxl as ox
            result['openpyxl'] = ox.__version__
            from openpyxl.styles import Font, PatternFill
            wb = ox.Workbook()
            ws = wb.active
            ws['A1'] = 'ok'
            ws['A1'].font = Font(bold=True, color='FFFFFF')
            ws['A1'].fill = PatternFill('solid', fgColor='1E4D78')
            result['workbook'] = 'OK'
        except Exception as e:
            result['workbook_error'] = str(e)
        resp = json.dumps(result, indent=2).encode()
        self.send_response(200)
        self._cors()
        self.send_header('Content-Type', 'application/json')
        self.end_headers()
        self.wfile.write(resp)

    def do_OPTIONS(self):
        self.send_response(200)
        self._cors()
        self.end_headers()

    def do_POST(self):
        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length)
        try:
            data = json.loads(body)
            b64 = gerar_planilha(data)
            resp = json.dumps({'b64': b64}).encode()
            self.send_response(200)
            self._cors()
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(resp)
        except Exception as e:
            err = json.dumps({'error': str(e)}).encode()
            self.send_response(500)
            self._cors()
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(err)

    def _cors(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')

    def log_message(self, format, *args):
        pass


# ─── Helpers de estilo ────────────────────────────────────────────────────────

def sf(bold=True, color="000000", size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def pf(color):
    return PatternFill("solid", fgColor=color)

def al(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def tb():
    t = Side(style="thin", color="000000")
    return Border(left=t, right=t, top=t, bottom=t)

def sc(ws, r, c, v, font=None, fill=None, align=None, border=None):
    cell = ws.cell(r, c, v)
    if font:   cell.font      = font
    if fill:   cell.fill      = fill
    if align:  cell.alignment = align
    if border: cell.border    = border
    return cell


# ─── Gerador principal ────────────────────────────────────────────────────────

def gerar_planilha(data):

    # Paleta de cores (sem prefixo FF — openpyxl aceita RRGGBB)
    AZUL_ESC  = "1E4D78"   # cabeçalhos principais
    AZUL_MED  = "3C78D8"   # cabeçalho Equipamentos
    AZUL_CLA  = "4A86E8"   # cabeçalho Usuários
    AZUL_DK   = "0B5394"   # cabeçalho "Tipo de Permissionamentos"
    AZUL_LINK = "0066CC"   # texto "Tipos de Permissões" e título col H
    BRANCO    = "FFFFFF"
    PRETO     = "000000"

    wb = openpyxl.Workbook()

    # ── ABA 1: Dados do Cliente ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Dados do Cliente"
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = Color(rgb="FF0066CC")

    # Larguras de coluna
    for col, w in [("A", 4.86), ("B", 19.29), ("C", 44.14),
                   ("D", 3.71),  ("E", 7.43),  ("F", 3.14),
                   ("G", 27.43), ("H", 35.86), ("I", 3.71)]:
        ws1.column_dimensions[col].width = w

    # Alturas de linha
    ws1.row_dimensions[1].height = 27
    ws1.row_dimensions[2].height = 12
    ws1.row_dimensions[3].height = 27
    for r in range(4, 36):
        ws1.row_dimensions[r].height = 15
    ws1.row_dimensions[21].height = 14.25  # linha levemente menor (igual ao modelo)

    # Linha 1 — título
    sc(ws1, 1, 2, "Clique ao Lado =>",
       font=sf(True, "FF0000", 12))
    sc(ws1, 1, 3, "Requisitos para implantação do Mobilemed",
       font=sf(False, PRETO, 12))

    # ── DADOS CADASTRAIS (B3:C3) ──
    ws1.merge_cells("B3:C3")
    sc(ws1, 3, 2, "DADOS CADASTRAIS",
       font=sf(True, BRANCO), fill=pf(AZUL_ESC), align=al("center"))

    campos_cliente = [
        ("Nome Fantasia:",  "nomefantasia"),
        ("Razão Social:",   "razaosocial"),
        ("CNPJ:",           "cnpj"),
        ("E-mail:",         "email"),
        ("Endereço:",       "endereco"),
        ("Bairro:",         "bairro"),
        ("Cidade:",         "cidade"),
        ("CEP:",            "cep"),
        ("Telefone:",       "telefone"),
    ]
    for i, (label, chave) in enumerate(campos_cliente, 4):
        sc(ws1, i, 2, label,               font=sf(True,  PRETO), align=al(), border=tb())
        sc(ws1, i, 3, data.get(chave, ""), font=sf(False, PRETO), align=al(), border=tb())

    # ── ROUTER DA MOBILEMED (G3:H3) ──
    ws1.merge_cells("G3:H3")
    sc(ws1, 3, 7, "ROUTER DA MOBILEMED",
       font=sf(True, BRANCO), fill=pf(AZUL_ESC), align=al())

    campos_router_mm = [
        ("IP:",                         "rmm_ip"),
        ("Porta:",                       "rmm_porta"),
        ("E-Title",                      "rmm_etitle"),
        ("AnyDesk / Team Viwer:",        "rmm_anydesk"),
        ("Senha:",                       "rmm_senha"),
        ("Login do Windows",             "rmm_login"),
        ("Senha do Usuário:",            "rmm_senhawin"),
        ("Localização Física do Router:", "rmm_local"),
        ("Observações:",                 "rmm_obs"),
    ]
    for i, (label, chave) in enumerate(campos_router_mm, 4):
        sc(ws1, i, 7, label,               font=sf(True,  PRETO), align=al(), border=tb())
        sc(ws1, i, 8, data.get(chave, ""), font=sf(False, PRETO), align=al(), border=tb())

    # ── INFORMAÇÕES DE SISTEMAS ADICIONAIS (B15:C15) ──
    ws1.merge_cells("B15:C15")
    sc(ws1, 15, 2, "INFORMAÇOES DE SISTEMAS ADCIONAIS",
       font=sf(True, BRANCO), fill=pf(AZUL_ESC), align=al())

    # Perguntas com merges B:C (linhas 16–25, pulando linhas pares para espaço)
    perguntas_sistemas = [
        (16, "1-Utiliza RIS e HIS? Qual?",                                       "rishis"),
        (18, "Se sim, terá integração com nosso sistema?",                       "integ1"),
        (20, "2-Tem consultórios internos ou externos, que acessam as imagens?", "consultorios"),
        (22, "3-Utiliza algum PACS atualmente? Qual?",                           "pacs"),
        (24, "Se sim, terá integração com nosso sistema?",                       "integpacs"),
    ]
    for row, label, chave in perguntas_sistemas:
        ws1.merge_cells(f"B{row}:C{row}")
        sc(ws1, row,     2, label,               font=sf(True,  PRETO), align=al(wrap=True), border=tb())
        ws1.merge_cells(f"B{row+1}:C{row+1}")
        sc(ws1, row + 1, 2, data.get(chave, ""), font=sf(False, PRETO), align=al(),          border=tb())

    # ── INFORMAÇÕES DO TI (G15:H15) ──
    ws1.merge_cells("G15:H15")
    sc(ws1, 15, 7, "INFORMAÇOES DO TI",
       font=sf(True, BRANCO), fill=pf(AZUL_ESC), align=al())

    campos_ti = [
        ("Responsável do T.I:", "ti_resp"),
        ("Fone TI:",             "ti_fone"),
        ("E-mail do TI:",        "ti_email"),
        ("Ti local ou tercerizado", "ti_tipo"),
        ("Observações:",         "ti_obs"),
    ]
    for i, (label, chave) in enumerate(campos_ti, 16):
        sc(ws1, i, 7, label,               font=sf(True,  PRETO), align=al(), border=tb())
        sc(ws1, i, 8, data.get(chave, ""), font=sf(False, PRETO), align=al(), border=tb())

    # ── ROUTER DA WORKLIST (G23:H23) ──
    ws1.merge_cells("G23:H23")
    sc(ws1, 23, 7, "ROUTER DA WORKLIST",
       font=sf(True, BRANCO), fill=pf(AZUL_ESC), align=al())

    campos_router_wl = [
        ("IP:",                  "rwl_ip"),
        ("Porta:",               "rwl_porta"),
        ("E-Title",              "rwl_etitle"),
        ("AnyDesk / Team Viwer:", "rwl_anydesk"),
        ("Senha:",               "rwl_senha"),
        ("Login do Windows",     "rwl_login"),
        ("Senha do Usuário:",    "rwl_senhawin"),
        ("Observações:",         "rwl_obs"),
    ]
    for i, (label, chave) in enumerate(campos_router_wl, 24):
        sc(ws1, i, 7, label,               font=sf(True,  PRETO), align=al(), border=tb())
        sc(ws1, i, 8, data.get(chave, ""), font=sf(False, PRETO), align=al(), border=tb())

    # ── ABA 2: Equipamentos ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Equipamentos")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = Color(rgb="FFFF9900")

    cabecalhos_eq = ["UNIDADE", "MODALIDADE", "MARCA", "MODELO",
                     "IP", "PORTA", "AETITLE", "Localização Física", "Observações"]
    larguras_eq   = [17, 15, 15, 20, 14, 10, 14, 25, 35]

    ws2.row_dimensions[1].height = 18
    for c, (h, w) in enumerate(zip(cabecalhos_eq, larguras_eq), 1):
        ws2.column_dimensions[get_column_letter(c)].width = w
        sc(ws2, 1, c, h,
           font=sf(True, BRANCO, 10), fill=pf(AZUL_MED),
           align=al("center"), border=tb())

    for i, eq in enumerate(data.get("equips", []), 2):
        ws2.row_dimensions[i].height = 15
        vals = [eq.get(k, "") for k in
                ("unidade", "modalidade", "marca", "modelo",
                 "ip", "porta", "aetitle", "local", "obs")]
        for c, v in enumerate(vals, 1):
            sc(ws2, i, c, v, font=sf(False, PRETO, 10),
               align=al(), border=tb())

    # ── ABA 3: Servidores ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Servidores")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = Color(rgb="FFFF9900")

    cabecalhos_sv = ["Unidade", "Serviços - router ou worklist",
                     "Sistema", "IP", "Acesso Externo SSH / Team",
                     "Usuário", "Senha"]
    larguras_sv   = [19.57, 25.71, 19.57, 14, 26.86, 19.57, 20]

    ws3.row_dimensions[1].height = 18
    for c, (h, w) in enumerate(zip(cabecalhos_sv, larguras_sv), 1):
        ws3.column_dimensions[get_column_letter(c)].width = w
        sc(ws3, 1, c, h,
           font=sf(True, BRANCO), fill=pf(AZUL_ESC),
           align=al("center"), border=tb())

    for i, sv in enumerate(data.get("servidores", []), 2):
        ws3.row_dimensions[i].height = 12.75
        vals = [sv.get(k, "") for k in
                ("unidade", "servico", "sistema", "ip",
                 "acesso", "usuario", "senha")]
        for c, v in enumerate(vals, 1):
            sc(ws3, i, c, v, font=sf(False, PRETO),
               align=al(), border=tb())

    # ── ABA 4: Usuários ───────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Usuários")
    ws4.sheet_view.showGridLines = False
    ws4.sheet_properties.tabColor = Color(rgb="FFFF0000")

    for col, w in [("A", 27.57), ("B", 30),    ("C", 19.29),
                   ("D", 17.57), ("E", 10.43),  ("F", 6.57),
                   ("G", 3),     ("H", 35.43)]:
        ws4.column_dimensions[col].width = w

    cabecalhos_usr = ["NOME COMPLETO", "EMAIL", "Telefone não Obrigatório",
                      "PERMISSAO", "CRM", "ESTADO"]
    ws4.row_dimensions[1].height = 18
    for c, h in enumerate(cabecalhos_usr, 1):
        sc(ws4, 1, c, h,
           font=sf(False, BRANCO, 10), fill=pf(AZUL_CLA),
           align=al("center"), border=tb())

    for i, u in enumerate(data.get("usuarios", []), 2):
        ws4.row_dimensions[i].height = 15
        vals = [u.get(k, "") for k in
                ("nome", "email", "telefone", "permissao", "crm", "estado")]
        for c, v in enumerate(vals, 1):
            sc(ws4, i, c, v, font=sf(False, PRETO, 10),
               align=al(), border=tb())

    # Coluna H — bloco de permissões (igual ao modelo)
    sc(ws4, 1, 8, "TIPO DE DE PERMISSIONAMENTOS",
       font=sf(True, BRANCO, 10), fill=pf(AZUL_DK), align=al("center"))

    tipos = ["MEDICO RADIOLOGISTA", "TECNICO", "DIGITADOR",
             "MEDICO SOLICITANTE", "GESTOR", "FINANCEIRO / RELATÓRIOS"]
    for i, p in enumerate(tipos, 2):
        sc(ws4, i, 8, p,
           font=sf(False, BRANCO, 10), fill=pf(AZUL_ESC), align=al())

    # Título das descrições (azul link, sem fundo)
    sc(ws4, 9, 8, "TIPOS DE PERMISSÕES EXISTENTES E SUAS DESCRIÇÕES",
       font=Font(bold=True, color=AZUL_LINK, size=10, name="Calibri"),
       align=al())

    descricoes = [
        (11, "TÉCNICO: Preparo de exames (estudo, lateralidade, nome paciente), "
             "conferência de imagens e consulta a pacientes."),
        (13, "SOMENTE LEITURA: Consulta lista de exames liberados, ou não, e laudos, "
             "pode ser usado por recepções, e demais pessoas que precisam apenas visualizar o laudo."),
        (15, "MÉDICO SOLICITANTE: Visualiza apenas os exames atribuídos à ele, cadastro individual."),
        (17, "MÉDICO EXECUTANTE: Assina os laudos, visualiza imagens."),
        (19, "GESTOR: Tem acesso a todas as ferramentas do portal (T.I e gerente do contrato);"),
        (21, "DIGITADOR (A): Acesso aos exames para auxiliar os médicos nos laudos."),
        (23, "FINANCEIRO: Acesso aos relatórios"),
    ]
    for row, texto in descricoes:
        ws4.row_dimensions[row].height = 30
        cell = ws4.cell(row, 8, texto)
        cell.font      = sf(False, PRETO, 10)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── ABA 5: Anotações Gerais ───────────────────────────────────────────────
    ws5 = wb.create_sheet("Anotações Gerais")
    ws5.sheet_view.showGridLines = False
    ws5.sheet_properties.tabColor = Color(rgb="FF000000")
    ws5.column_dimensions["A"].width = 100

    for i, linha in enumerate((data.get("anotacoes", "") or "").split("\n"), 1):
        ws5.cell(i, 1, linha).font = sf(False, PRETO)

    # ── Serializar e retornar base64 ──────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode()